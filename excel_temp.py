'''
World Temperature Database Chinese Cities Data Processing Script
Author: Hashim-Jones, Jake (21/09/2017)
Version 1.0.

This script creates a connection to the database created by db_create.py. It then queries the database for average annual temperature data from all major cities in China. It then processes this data and exports it to a spreadshxeet in an excel workbook ('World Temperatures.xlsx') and creates a line chart (also saved in the spreadsheet).

See readme for more details.
'''

########################################
## Importing and Creating Definitions ##
########################################

print("Initializing.\n")
import openpyxl
from openpyxl.chart import (LineChart, Reference)
from openpyxl.styles import (Font, Alignment, PatternFill, Color, Border, Side)
import sqlite3
from os.path import isfile
import datetime

def yesNoInput(prompt=""):
    while True:
        answer = input(prompt).upper() # Get and capitalize user input.
        if answer == "Y" or answer == "N":
            break
        print("Error! Please enter 'Y' or 'N' (case-insensitive)") # Reprompt otherwise.
    return answer == "Y" # Return True if the user input 'Y' (meaning yes), or False if the user input 'N'.

################################
## Create Database Connection ##
################################

if not isfile("Temperature_Data.db"): # Check that database file exists...
    print("Error, 'Temperature_Data.db' does not exist. Please run 'db_create.py' first.") # Prompt to run the creation script if it does not.
    exit(0)

dbConnection = sqlite3.connect("Temperature_Data.db") # Open connection to the (existing) database.
print("\nOpening 'Temperature_Data.db'")
print("Connected to database.", datetime.datetime.now(),"\nSqlite version:", sqlite3.sqlite_version, "\n")
dbCursor = dbConnection.cursor() # Create cursor object.

print("Checking Tables...\n")
existingTables = dbCursor.execute("Select Name From sqlite_master Where type = 'table';").fetchall() # Obtaining table names that exist in the database.
existingTables = [name[0] for name in existingTables] # Extracting table names from the database output.
missingTables = [name for name in ['Country', 'MajorCity', 'State'] if name not in existingTables] # Compile list of 'expected' tables in the database that are NOT present.

if len(missingTables) >0: # This branch is executed if the database is mising one of the essential 3 tables.
    for name in missingTables:
        print("'{}' Table is missing.".format(name)) # Alert user that table(s) are missing.
    print("\nError, 'Temperature_Data.db has incomplete data. Please run 'db_create.py' to ensure all appropriate data is available for the program.\n") # Prompt user to run the creation script.
    print("Disconnecting from the database...")
    dbConnection.close() # Close the database connection.
    print("Disconnected from the database.", datetime.datetime.now())
    exit(0) # Terminate the program.
else: # Everything is fine and the program continues.
    print("Success.Check Complete. Database has all appropriate data.\n\n")

###############################
## Creating/Opening Workbook ##
###############################

if isfile("World Temperature.xlsx"): # Check if workbook already exists. If it does already exist this branch is executed.
    if not yesNoInput("Warning! Workbook 'World Temperature.xlsx' already exists. Continuing modify this workbook. Do you wish to continue (Y/N)?"): # If workbook does exist, warn user and give option to abort.
        exit(0)
    else:
        print("\nOpening Workbook...")
        worldTempWB = openpyxl.load_workbook("World Temperature.xlsx") # Load existing workbook.
        print("Success.\n")
        print("Checking existing sheets...\n")
        sheets = worldTempWB.get_sheet_names() # Retrieve a list of sheet names from the existing workbook.
        if "Temperature by City" in sheets: # Determine if the sheet 'Temperature by City' sheet already exists. The following branch is executed if it does exist.
            if not yesNoInput("Warning! 'Temperature by City' is already in the workbook sheets. Continuing will replace the data in this sheet. Do you wish to continue (Y/N)? "): # Warn user and give the option to abort the script.
                exit(0)
            else:
                print("\nRemoving sheet 'Temperature by City'...")
                sheetToRemove = worldTempWB.get_sheet_by_name("Temperature by City") # Obtain existing 'Temperature by City' sheet.
                worldTempWB.remove(sheetToRemove) # Delete existing 'Temperature by City' sheet.
                print("Success.\n")
        else: # If 'Temperature by City' sheet does not already exist, continue the program.
            print("No conflicting sheets.\n")
else: # Workbook does not exist already.
    print("Creating new workbook...")
    worldTempWB = openpyxl.Workbook() # Create new workbook.
    print("Success.\n")
    print("Cleaning new workbook file...")
    defaultSheet = worldTempWB.get_sheet_by_name("Sheet") # Obtain default sheet.
    worldTempWB.remove(defaultSheet) # Remove default sheet.
    print("Success. Removed all default material.\n")

print("Creating Worksheet 'Temperature by City'...")
worldTempWS = worldTempWB.create_sheet("Temperature by City") # Create new worksheet called 'Temperature by City'.
print("Success.\n\n")

########################
## Query the Database ##
########################

print("Obtaining temperature data from major Chinese cities...\n")
print("Warning! Some data may missing. These will not be counted in the calculated averages.\n")
data=dbConnection.execute('''
SELECT SUBSTR(date, 0, 5) As Year, City, AVG(AverageTemperature) 
FROM MajorCity
WHERE Country='China'
GROUP BY Year, City
ORDER BY Year, City;
''').fetchall() # Query the database to retrieve city temperature data for all cities in China.
print("Data retrieved.\n\n")

###############################
## Close Database Connection ##
###############################

print("Disconnecting from the database...")
dbConnection.close() # Close database connection.
print("Disconnected from the database.", datetime.datetime.now(), "\n\n")

###########################
## Add Data to Worksheet ##
###########################

cities={}
records={}
print("Compiling data...\n")
for record in data:
    records[record[0]] = records.get(record[0], []) + [(record[1], record[2])] # Adds records if to dictionary (key is the year and the value is a tuple of city and average temperature).
    cities[record[1]] = cities.get(record[1], 0) + 1 # Generates a dictionary of ALL city names (for buffering later on).
print("Warning! Some of the data may be missing for average annual temperature. These values will be added as blank cells.")
# City names will automatically be in the order of the number of records due to the way that the sql query returns data (and the processing above).
citylist = list(cities.keys())

# The following loop essentially adds an empty value for a city (with a temperature of None), if that city was not returned from the original query (ie. there wasn't even an empty record for that city).
for record in records:
    newRecordEntries = [(city, None) for city in citylist] # Generates a list of 'placeholder' records (this gives appropriate empty space buffers when adding data to the excel sheet).
    newRecordKeys={}
    for item in newRecordEntries: # Adds records to a dictionary as keys.
        newRecordKeys[item]=True
    for item in records[record]: # Replaces placeholder records if actual data record exists.
        if (item[0], None) in newRecordKeys.keys():
            del newRecordKeys[(item[0], None)] # Remove placeholder record
            newRecordKeys[(item[0], item[1])] = True # Add actual data in its place.
    records[record] = sorted(list(newRecordKeys.keys())) # Sort the records by city.
print("Complete.\n")

# The following loop generates the rows and adds them to the spreadsheet.
print("Generating rows and adding to spreadsheet...")
worldTempWS.append(['Year'] + sorted(list(cities.keys()))) # Add header row to spreadsheet
for year in records:
     temperatures = [record[1] for record in records[year]] # Compile row for spreadsheet.
     worldTempWS.append([year] + temperatures) # Add row to spreadsheet.
     print("     Added temperature data for {}".format(year))
print("Success. All data has been added to the spreadsheet.\n\n")

#####################################
## Formatting the Spreadsheet Data ##
#####################################

print("Formatting the spreadsheet...")
# The following section defines various formatting styles for the spreadsheet.
titleFont = Font(size=14, bold=True)
headingFont=Font(bold=True)
centeredMissingValue=Alignment(horizontal='center')
colourTopRow = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('69addb'))
colourFirstColumn = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('95c3e2'))
colourDataCell = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('c9e8fc'))
bottomBorderOnly = Border(bottom = Side(border_style='thick', color='FF000000'))
rightBorderOnly = Border(right = Side(border_style='thick', color='FF000000'))
topBorderOnly  = Border(top = Side(border_style='thick', color='FF000000'))
leftBorderOnly  = Border(left = Side(border_style='thick', color='FF000000'))
# End of the styles definition section.

# The following section applies formatting to the spreadsheet.
worldTempWS.freeze_panes = 'B2' # Frezzes first row and first column in the spreadsheet.
worldTempWS.column_dimensions['A'].width = 25 # Manually set the width of the first column.
for row in worldTempWS.iter_rows(max_row=worldTempWS.max_row, max_col=worldTempWS.max_column, min_row=1, min_col=1): #Iterate through spreadsheet rows
    for cell in row:
        if cell.row == 1: # Affects the first row.
            cell.font = titleFont # Adjusts font of the first (title) row.
            cell.fill = colourTopRow # Adjusts the background colour of the first row.
            cell.border = bottomBorderOnly # Adds bottom border to the first row.
        if cell.column == 'A' and cell.row != 1: # Affects the first column except the cell 'A1'.
            cell.font = headingFont # Adjust font of first column.
            cell.fill = colourFirstColumn # Adjust background colour of first column.
            cell.border = rightBorderOnly # Add right border to first column.
        if cell.column != 'A' and cell.row != 1: # Affect all valud cells except those in the first row or column (data cells).
            cell.fill = colourDataCell # Change background colour of data cells.
for row in worldTempWS['A196:Q196']: # Affects the row below the last data row.
    for cell in row:
        cell.border = topBorderOnly # Adds border to the bottom of the data area.
for row in worldTempWS['R1:R195']: # Affects the column below the last data column.
    for cell in row:
        cell.border = leftBorderOnly # Adds border to the right side od the data area.
print("Done.\n\n")
#End of the formatting section


#########################
## Generate Line Chart ##
#########################

# Constructing an empty line chart.
print("Constructing line chart...")
print("     Generating line chart.")
chart = LineChart() # Create LineChart object.
print("     Formatting line chart.")
chart.title = "Average Annual Temperature In Major Chinese Cities" # Set chart title.
chart.y_axis.title = u'Average Yearly Temperature (\xb0C)' # Set y-axis title.
chart.x_axis.title = "Year" # Set x-axis title.
chart.height *= 2 # Increase the height of the chart by a factor of 2.
chart.width *= 2 # Increase the width of the chart by a factor of 2.

# Adding data to the line chart (as references).
print("     Adding data to line chart.")
chartData=Reference(worldTempWS, min_col=2, min_row=1, max_col=((worldTempWS.max_column)-1), max_row=((worldTempWS.max_row)-1)) # Create Reference object.
chart.add_data(chartData, titles_from_data=True) # Add Reference object to chart.
print("     Adding x-axis values to line chart.")
years = Reference(worldTempWS, min_col=1, max_col=1,min_row=2, max_row=((worldTempWS.max_row)-1)) # Create Reference object.
chart.set_categories(years) # Use reference object to categorize chart data along the x-axis (ie. use data as x-axis labels).

print("     Adding line chart to spreadsheet.")
worldTempWS.add_chart(chart,"B{}".format((worldTempWS.max_row)+2)) # Add chart to spreadsheet.
print("Success.\n\n")

###################
## Save Workbook ##
###################

print("Saving changes...")
if isfile("World Temperature.xlsx"): # If workbook already exists...
    if yesNoInput("Are you sure you would like to save changes to 'World Temperature.xlsx'? This action cannot be undone (Y/N). "): # Give user the option to save changes.
        worldTempWB.save("World Temperature.xlsx") # Save changes.
        print("\nChanges have been saved.")
    else:
        print("\nChanges have not been saved.")
else: # Workbook does not already exist.
    if yesNoInput("Would you like to save 'World Temperature.xlsx' (Y/N)? "): # Give user the option to save a new file.
        worldTempWB.save("World Temperature.xlsx") # Save new file.
        print("\nWorkbook has been saved.")
    else:
        print("\nWorkbook has not been saved.")
