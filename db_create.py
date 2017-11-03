'''
World Temperature Database Creation Script
Author: Hashim-Jones, Jake (21/09/2017)
Version 1.0.

This script creates a database (containing three tables) and imports data from three pre-created excel workbooks (file names are outlined in readme.txt). The database is stored locally (as Temperature_Data.db) and runs via the sqlite DBMS. The database is also indexed to provide optimum speeds when running the rest of the scripts in the set.

See readme for more details.
'''

########################################
## Importing and Creating Definitions ##
########################################

print("Initializing")
import sqlite3
import openpyxl
from os.path import isfile
import datetime

def yesNoInput(prompt=""):
    while True:
        answer = input(prompt).upper() # Get and capitalize user input.
        if answer == "Y" or answer == "N":
            break
        print("Error! Please enter 'Y' or 'N' (case-insensitive)") # Reprompt otherwise.
    return answer == "Y" # Return True if the user input 'Y' (meaning yes), or False if the user input 'N'.

################################
## Create Database Connection ##
################################

databaseAlreadyExists=False
if isfile("Temperature_Data.db"): # Check for if the file exists already or not.
    databaseAlreadyExists = yesNoInput("Warning! 'Temperature_Data.py' already exists. Would you like to continue (Y/N)? ") # If the file does exist, give user the opportunity to abort the script.
    if not databaseAlreadyExists:
        exit(0)

dbConnection = sqlite3.connect("Temperature_Data.db") # Open connection to a new database file (or an existing one if it already exists).
print("\nOpening 'Temperature_Data.db'")
print("Connected to database.", datetime.datetime.now(),"\nSqlite version:", sqlite3.sqlite_version, "\n")
dbCursor = dbConnection.cursor() # Create cursor object.

if databaseAlreadyExists: # Only execute this branch if the database already exists...
    print("Checking Tables\n")
    existingTables = dbCursor.execute("Select Name From sqlite_master Where type = 'table';").fetchall() # Retrieve all table names from the database.
    if len(existingTables) != 0:
        print("The following tables already exist...")
        for table in existingTables:
            print("     {}".format(table[0])) # Display list of existing table names to user.
        print("\nContinuing will override these tables. This action cannot be undone.")
        if yesNoInput("Are you sure you would like to continue (Y/N)? "): # Gives user the option to abort the script.
            print("\n")
            for table in existingTables:
                print("Dropping Table - {}".format(table[0])) 
                dbCursor.execute('Drop Table "{}";'.format(table[0])) # Drop ALL tables in the database already.
            if len(dbCursor.execute("Select Name From sqlite_master Where type = 'table';").fetchall()) == 0: # Check that table dropping has been successful.
                print("\nTables Successfully Dropped.\n")
        else:
            print("Disconnecting from the database...")
            dbConnection.close() # Close the database connection.
            print("Disconnected from the database.", datetime.datetime.now())
            exit(0)
    else:
        print("'Temperature_Data.db' does not contain any tables.Continuing program.\n\n")

###################################
## Import Data From Spreadsheets ##
###################################

print("Importing data from 'GlobalLandTemperaturesByCountry.xlsx'")
try:
    temperatureByCountryWB = openpyxl.load_workbook("GlobalLandTemperaturesByCountry.xlsx") # Load excel workbook into memory.
except:
    print("Error, 'GlobalLandTemperaturesByCountry.xlsx' not found.") # Error handling for if workbook not found.
    exit(0)
getSheets = temperatureByCountryWB.get_sheet_names() # Get sheet names from the loaded workbook.
if len(getSheets) > 1:
    print("Error, there are too many sheets in the workbook. Getting data from the first sheet only.") # Only takes the first sheet if there are multiple sheets in the workbook.
elif len(getSheets) == 0:
    print("Error, there are no sheets in this workbook.") # Error for if workbook is empty.
    exit(0)
temperatureByCountryWS = temperatureByCountryWB.get_sheet_by_name(getSheets[0]) # Assign workbook sheet to variable.
print("Success.\n")


print("Importing data from 'GlobalLandTemperaturesByMajorCity.xlsx'")
try:
    temperatureByMajorCityWB = openpyxl.load_workbook("GlobalLandTemperaturesByMajorCity.xlsx") # Load excel workbook into memory.
except:
    print("Error, 'GlobalLandTemperaturesByMajorCity.xlsx' not found.") # Error handling for if workbook not found.
    exit(0)
getSheets = temperatureByMajorCityWB.get_sheet_names() # Get sheet names from the loaded workbook.
if len(getSheets) > 1:
    print("Error, there are too many sheets in the workbook. Getting data from the first sheet only.") # Only takes the first sheet if there are multiple sheets in the workbook.
elif len(getSheets) == 0:
    print("Error, there are no sheets in this workbook.") # Error for if workbook is empty.
    exit(0)
temperatureByMajorCityWS = temperatureByMajorCityWB.get_sheet_by_name(getSheets[0]) # Assign workbook sheet to variable.
print("Success.\n")


print("Importing data from 'GlobalLandTemperaturesByState.xlsx'")
try:
    temperatureByStateWB = openpyxl.load_workbook("GlobalLandTemperaturesByState.xlsx") # Load excel workbook into memory.
except:
    print("Error, 'GlobalLandTemperaturesByState.xlsx' not found.") # Error handling for if workbook not found.
    exit(0)
getSheets = temperatureByStateWB.get_sheet_names() # Get sheet names from the loaded workbook.
if len(getSheets) > 1:
    print("Error, there are too many sheets in the workbook. Getting data from the first sheet only.") # Only takes the first sheet if there are multiple sheets in the workbook.
elif len(getSheets) == 0:
    print("Error, there are no sheets in this workbook.") # Error for if workbook is empty.
    exit(0)
temperatureByStateWS = temperatureByStateWB.get_sheet_by_name(getSheets[0]) # Assign workbook sheet to variable.
print("Success.\n")

############################
## Create Database Tables ##
############################

#Create table for data in 'GlobalLandTemperaturesByCountry.xlsx'.
titles=[]
for row in temperatureByCountryWS.iter_rows(max_row=temperatureByCountryWS.min_row, max_col=temperatureByCountryWS.max_column):
    for cell in row:
        titles.append(cell.value) # Import attribute names directly from spreadsheet headings.
        
temperatureByCountryTable = """
Create Table Country(
    {attr1} Date,
    {attr2} Decimal,
    {attr3} Decimal,
    {attr4} Varchar2(30),
    CONSTRAINT country_PK Primary Key ({attr1},{attr4})
);
""".format(attr1=titles[0], attr2=titles[1], attr3=titles[2], attr4=titles[3]) # Schema for the 'Country' table.

dbCursor.execute(temperatureByCountryTable) # Add 'Country' table to the database.
print("\n'Country' Table has been created!")

#Create table for data in 'GlobalLandTemperaturesByMajorCity.xlsx'
titles=[]
for row in temperatureByMajorCityWS.iter_rows(max_row=temperatureByMajorCityWS.min_row, max_col=temperatureByMajorCityWS.max_column):
    for cell in row:
        titles.append(cell.value) # Import attribute names directly from spreadsheet headings.
        
temperatureByMajorCityTable = """
Create Table MajorCity(
    {attr1} Date,
    {attr2} Decimal,
    {attr3} Decimal,
    {attr4} Varchar2(30),
    {attr5} Varchar2(30),
    {attr6} Varchar2(9),
    {attr7} Varchar2(9),
    CONSTRAINT majorcity_PK Primary Key ({attr1},{attr4},{attr5})
);
""".format(attr1=titles[0], attr2=titles[1], attr3=titles[2], attr4=titles[3], attr5=titles[4], attr6=titles[5], attr7=titles[6]) # Schema for the 'MajorCity' table.

dbCursor.execute(temperatureByMajorCityTable) # Add 'MajorCity' table to the database.
print("'MajorCity' Table has been created!")


#Create table for data in 'GlobalLandTemperaturesByState.xlsx'
titles = []
for row in temperatureByStateWS.iter_rows(max_row=temperatureByStateWS.min_row,
                                              max_col=temperatureByStateWS.max_column):
    for cell in row:
        titles.append(cell.value) # Import attribute names directly from spreadsheet headings.

temperatureByStateTable = """
Create Table State(
    {attr1} Date,
    {attr2} Decimal,
    {attr3} Decimal,
    {attr4} Varchar2(30),
    {attr5} Varchar2(30),
    CONSTRAINT state_PK Primary Key ({attr1},{attr4},{attr5})
);
""".format(attr1=titles[0], attr2=titles[1], attr3=titles[2], attr4=titles[3], attr5=titles[4]) # Schema for the 'State' table.

dbCursor.execute(temperatureByStateTable) # Add 'State' table to the database.
print("'State' Table has been created!\n")


##########################################
## Add Excel Data From Each Spreadsheet ##
##########################################

print("Addding data to 'Country' table...")
for line in temperatureByCountryWS.iter_rows(min_row=2, max_row=temperatureByCountryWS.max_row, max_col=temperatureByCountryWS.max_column):
    row=[cell.value if cell.value != None else 'Null' for cell in line ] # Add null for missing data values.
    dbCursor.execute('Insert Into Country Values ("{Date}",{AverageTemperature},{AverageTemperatureUncertainty},"{Country}");'.format(Date=row[0], AverageTemperature=row[1], AverageTemperatureUncertainty=row[2], Country=row[3])) # Add each row to the 'Country' table in the database
print("Done.\n")

print("Addding data to 'MajorCity' table...")
for line in temperatureByMajorCityWS.iter_rows(min_row=2, max_row=temperatureByMajorCityWS.max_row, max_col=temperatureByMajorCityWS.max_column):
    row=[cell.value if cell.value != None else 'Null' for cell in line ] # Add null for missing data values.
    dbCursor.execute('Insert Into MajorCity Values ("{Date}",{AverageTemperature},{AverageTemperatureUncertainty},"{City}", "{Country}", "{Latitude}", "{Longitude}");'.format(Date=row[0], AverageTemperature=row[1], AverageTemperatureUncertainty=row[2], City=row[3], Country=row[4], Latitude=row[5], Longitude=row[6])) # Add each row to the 'MajorCity' table in the database.
print("Done.\n")

print("Addding data to 'State' table...")
for line in temperatureByStateWS.iter_rows(min_row=2, max_row=temperatureByStateWS.max_row, max_col=temperatureByStateWS.max_column):
    row=[cell.value if cell.value != None else 'Null' for cell in line ] # Add null for missing data values.
    dbCursor.execute('Insert Into State Values ("{Date}",{AverageTemperature},{AverageTemperatureUncertainty},"{State}", "{Country}");'.format(Date=row[0], AverageTemperature=row[1], AverageTemperatureUncertainty=row[2], State=row[3], Country=row[4])) # Add each row to the 'State' table in the database.
print("Done.\n")

###########################################
## Index Database to Improve Performance ##
###########################################

#Optimizes the database for the queries performed by the set of scripts. Indexed attributes are the ones that appear in the 'WHERE' clause of SQL queries. NOTE: This is done AFTER the data has been insesrted as indexing will slow down future data insertion.
print("Optimizing database by performing indexing...")
dbCursor.execute("CREATE INDEX i_state_country ON State(Country);")
dbCursor.execute("CREATE INDEX i_state_state ON State(State);")
dbCursor.execute("CREATE INDEX i_state_date ON State(date);")
dbCursor.execute("CREATE INDEX i_majorcity_country ON MajorCity(Country);")
dbCursor.execute("CREATE INDEX i_majorcity_latitude ON MajorCity(Latitude);")
dbCursor.execute("CREATE INDEX i_majorcity_date ON MajorCity(date);")
dbCursor.execute("CREATE INDEX i_country_country ON Country(country);")
print("Done.")

##################################################
## Commit Database Changes and Close Connection ##
##################################################

print("\nCommiting changes to the database...")
dbConnection.commit() # Commit changes to the database.
print("Success.\n")
print("Disconnecting from the database...")
dbConnection.close() # Close the connection.
print("Disconnected from the database.",datetime.datetime.now())
