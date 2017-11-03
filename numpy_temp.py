'''
World Temperature Database Australian State Temperature Data Processing Script
Author: Hashim-Jones, Jake (21/09/2017)
Version 1.0.

This script creates a connection to the database created by db_create.py. It then queries the database for average annual temperature data states in Australia (and for the nation itself) AND calculates the differences between each state and the national data (for each year). It then processes this data and exports it to a spreadshxeet in an excel workbook ('World Temperatures.xlsx'). This data is plotted (using matplotlib) and shown on screen in a seperate new window.

See readme for more details.
'''

########################################
## Importing and Creating Definitions ##
########################################

print("Initializing.\n")
import openpyxl
from openpyxl.styles import (Font, Alignment, PatternFill, Color, Border, Side)
import sqlite3
from os.path import isfile
import datetime
import numpy
import matplotlib.pyplot as plt

def yesNoInput(prompt=""):
    while True:
        answer = input(prompt).upper() # Get and capitalize user input.
        if answer == "Y" or answer == "N":
            break
        print("Error! Please enter 'Y' or 'N' (case-insensitive)") # Reprompt otherwise.
    return answer == "Y" # Return True if the user input 'Y' (meaning yes), or False if the user input 'N'.

################################
## Create Database Connection ##
################################

if not isfile("Temperature_Data.db"): # Check that database file exists...
    print("Error, 'Temperature_Data.db' does not exist. Please run 'db_create.py' first.") # Prompt to run the creation script if it does not.
    exit(0)

dbConnection = sqlite3.connect("Temperature_Data.db") # Open connection to the (existing) database.
print("\nOpening 'Temperature_Data.db'")
print("Connected to database.", datetime.datetime.now(),"\nSqlite version:", sqlite3.sqlite_version, "\n")
dbCursor = dbConnection.cursor() # Create cursor object.

print("Checking Tables...\n")
existingTables = dbCursor.execute("Select Name From sqlite_master Where type = 'table';").fetchall() # Obtaining table names that exist in the database.
existingTables = [name[0] for name in existingTables] # Extracting table names from the database output.
missingTables = [name for name in ['Country', 'MajorCity', 'State'] if name not in existingTables] # Compile list of 'expected' tables in the database that are NOT present.

if len(missingTables) >0: # This branch is executed if the database is mising one of the essential 3 tables.
    for name in missingTables:
        print("'{}' Table is missing.".format(name)) # Alert user that table(s) are missing.
    print("\nError, 'Temperature_Data.db has incomplete data. Please run 'db_create.py' to ensure all appropriate data is available for the program.\n") # Prompt user to run the creation script.
    print("Disconnecting from the database...")
    dbConnection.close() # Close the database connection.
    print("Disconnected from the database.", datetime.datetime.now())
    exit(0) # Terminate the program.
else: # Everything is fine and the program continues.
    print("Success.Check Complete. Database has all appropriate data.\n\n")

###############################
## Creating/Opening Workbook ##
###############################

if isfile("World Temperature.xlsx"): # Check whether or not workbook already exists. This branch is executed if it does.
    if not yesNoInput("Warning! Workbook 'World Temperature.xlsx' already exists. Continuing modify this workbook. Do you wish to continue (Y/N)?"): # Warn user and give option to abort the script.
        exit(0)
    else:
        print("\nOpening Workbook...")
        worldTempWB = openpyxl.load_workbook("World Temperature.xlsx") # Open existing workbook.
        print("Success.\n")
        print("Checking existing sheets...\n")
        sheets = worldTempWB.get_sheet_names() # Obtain a list of sheet names from the workbook.
        if "Comparison" in sheets: # Look for the 'Comparison' sheet in the sheet names. This branch is executed if the sheet already exists.
            if not yesNoInput("Warning! 'Comparison' is already in the workbook sheets. Continuing will replace the data in this sheet. Do you wish to continue (Y/N)? "): # Warn user and give the option to abort the script.
                exit(0)
            else:
                print("\nRemoving sheet 'Comparison'...")
                sheetToRemove = worldTempWB.get_sheet_by_name("Comparison") # Obtain existing 'Comparison' sheet.
                worldTempWB.remove(sheetToRemove) # Delete existing 'Comparison' sheet.
                print("Success.\n")
        else: # Continue program if 'Comparison' sheet does not already exist.
            print("No conflicting sheets.\n")
else: # Workbook does not exist already.
    print("Creating new workbook...")
    worldTempWB = openpyxl.Workbook() # Create new workbook.
    print("Success.\n")
    print("Cleaning new workbook file...")
    defaultSheet = worldTempWB.get_sheet_by_name("Sheet") # Obtain default sheet.
    worldTempWB.remove(defaultSheet) # Remove default sheet.
    print("Success. Removed all default material.\n")

print("Creating Worksheet 'Comparison'...")
worldTempWS = worldTempWB.create_sheet("Comparison") # Create new worksheet called 'Comparison'.
print("Success.\n\n")

########################
## Query the Database ##
########################

print("Obtaining temperature data from Australian states...")


statenames=[entry[0] for entry in dbConnection.execute('''
SELECT DISTINCT State 
From State
WHere Country='Australia';
''').fetchall()] # Query the database and obtain a list of all possible state names in Australia.
print("     Retrieved state names.")

stateData={}
years={}
for state in statenames: # This loop retrieves data from individual states based on the state names retrieved above.
    result = dbConnection.execute('''
    SELECT CAST(SUBSTR(date, 0, 5) As INTEGER) As Year, AVG(AverageTemperature) 
    FROM State
    WHERE Country='Australia'
        AND State='{}'
    GROUP BY Year, State
    ORDER BY Year, State;
    '''.format(state)).fetchall() # Query the database to retrieve relevant data for each state in Australia (from the 'State' table).
    for record in result: # Notes all the years found across all data.
        years[record[0]]=True
    stateData[state]=result # Adds each data set to a dictionary as a list (the state names are the keys).
    print("     Retrieved data for {}.".format(state))

result=dbConnection.execute('''
SELECT CAST(SUBSTR(date, 0, 5) As INTEGER) As Year, AVG(AverageTemperature) 
FROM Country
WHERE Country='Australia'
GROUP BY Year;
''').fetchall() # Queries the database to retrieve all national average data (from the 'Country' table).
print("     Retrieved national temperature data.")
stateData['Australia'] = result # Adds retrieved data to a new entry in the data dictionary under the 'Australia' key.
for record in result:  # Notes all the years found across all data.
    years[record[0]] = True
print("Success.\n\n")

###############################
## Close Database Connection ##
###############################

print("Disconnecting from the database...")
dbConnection.close() # Disconnect from the database.
print("Disconnected from the database.",datetime.datetime.now(), "\n\n")

#####################
## Processing Data ##
#####################
print("Processing data...")

# The following loop adds None values for years where the data is missing (even in cases where the year was not returned by the query).
for record in stateData:
    newRecordEntries = [(year, None) for year in years] # Generates blank list of tuples containing all possible years retrieved with null values. These are simply placeholders.
    newRecordKeys={}
    for item in newRecordEntries: # Compiles dictionary containing all placeholder tuples generated previously
        newRecordKeys[item]=True
    for item in stateData[record]: # This loop replaces place holder tuples in the dictionary with data from the database if it is present.
        if (item[0], None) in newRecordKeys.keys():
            del newRecordKeys[(item[0], None)] # Remove placeholder tuple.
            newRecordKeys[(item[0], item[1])] = True # Add data tuple from database.
    newRecord = sorted(list(newRecordKeys.keys())) # Obtain list from dictionary keys (ordered by year).
    arrayData = [temperature[1] for temperature in newRecord] # Compile a list of temperatures without year (still ordered by year).
    stateData[record] = numpy.array(arrayData, dtype=float) # Convert list to an ndarray.

differences={}
for state in stateData:
    if state=='Australia':
        continue
    differences[state]=stateData[state]-stateData['Australia'] # Calculate a new set of data which are the differences between each state data set and the national data set (each result is added to a new dictionary.
print("Success.\n\n")

########################
## Generate Data Plot ##
########################

print("Generating plot...")
y = list(years.keys()) # Defining the x-axis

# The following creates a new figure and adds each data set (from differences) as a new subplot on the figure.
plt.figure(1)
for state in differences:
    plt.subplot(111) # Add new subplot.
    plt.plot(y, differences[state], linestyle=' ', marker='.', label=state) # Plot each state's difference data.
print("Success.\n")

#The following formats the figure to make it more presentable aesthetically.
plt.axhline(y=0, color='k', linestyle='-') # Flat line for y=0.
print("Adding titles and legends to plot...")
plt.legend()
plt.grid(True, which='both', linestyle='--')
plt.title("Differences in State Average Annual Temperature With National Average Annual Temperature")
plt.xlabel("Year")
plt.ylabel(u"Difference Between State and National Average Annual Temperature (\xb0C)")

print("Success.\n")

print("Opening plot...")
plt.show() # Outputs the plot in a seperate window (unless in ipython console).
print("Plot closed.\n\n")

# The following block of commented code creates and displays a second plot of the average annual national data for Australia over the years. This can be uncommented to produce this plot in addition to the one generated above.

# plt.plot(y, stateData['Australia'])
# print("Adding titles and legends to plot...")
# plt.grid(True, which='both', linestyle='--')
# plt.title("National Average Temperature Between 1852 and 2013")
# plt.xlabel("Year")
# plt.ylabel(u"National Average Annual Temperature (\xb0C)")
# plt.show()


#############################
## Add Data to Spreadsheet ##
#############################

# The following appends the original data sets and calculated ones to the spreadsheet appropriately (row by row).
print("Adding data to spreadsheet...")
worldTempWS.append(['Year'] + y) # Add first row (title row) containing all possible years in the data.
worldTempWS.append(["Australia"] + [temp if not numpy.isnan(temp) else '-' for temp in list(stateData['Australia'])]) # Add national temperature data with '-' if the value is nan (not a number).
worldTempWS.append([None]) # Add empty row.
worldTempWS.append(['Individual State Temperature Data'])
for state in stateData:
    worldTempWS.append([state] + [temp if not numpy.isnan(temp) else '-' for temp in list(stateData[state])]) # Add state temperature data with '-' if the value is nan (not a number). Loop iterates through the states.
worldTempWS.append([None]) # Add empty row.
worldTempWS.append(["Difference Between State and National Avarage Temperature"])
for state in differences:
    worldTempWS.append([state] + [temp if not numpy.isnan(temp) else '-' for temp in list(differences[state])]) # Add state difference data with '-' if the value is nan (not a number). Loop iterates through the states.
print("Success.\n")

#####################################
## Formatting the Spreadsheet Data ##
#####################################

print("Formatting the spreadsheet...")
# The following section defines various formatting styles for the spreadsheet.
titleFont = Font(size=14, bold=True)
headingFont=Font(bold=True)
centeredMissingValue=Alignment(horizontal='center')
colourTopRow = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('69addb'))
colourFirstColumn = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('95c3e2'))
colourDataCell = PatternFill(patternType='solid', fill_type='solid', fgColor=Color('c9e8fc'))
bottomBorderOnly = Border(bottom = Side(border_style='thick', color='FF000000'))
rightBorderOnly = Border(right = Side(border_style='thick', color='FF000000'))
topBorderOnly  = Border(top = Side(border_style='thick', color='FF000000'))
leftBorderOnly  = Border(left = Side(border_style='thick', color='FF000000'))
# End of the styles definition section.

# The following section applies formatting to the spreadsheet.
worldTempWS.freeze_panes = 'B2' # Frezzes first row and first column in the spreadsheet.
worldTempWS.column_dimensions['A'].width = 25 # Manually set the width of the first column.
for row in worldTempWS.iter_rows(max_row=worldTempWS.max_row, max_col=worldTempWS.max_column, min_row=1, min_col=1): #Iterate through spreadsheet rows
    for cell in row:
        if cell.value == '-':
            cell.alignment = centeredMissingValue # Centres '-' in cells that do not contain data value.
        if cell.row == 1: # Affects the first row.
            cell.font = titleFont # Adjusts font of the first (title) row.
            cell.fill = colourTopRow # Adjusts the background colour of the first row.
            cell.border = bottomBorderOnly # Adds bottom border to the first row.
        if cell.column == 'A' and cell.row != 1: # Affects the first column except the cell 'A1'.
            cell.font = headingFont # Adjust font of first column.
            cell.fill = colourFirstColumn # Adjust background colour of first column.
            cell.border = rightBorderOnly # Add right border to first column.
        if cell.column != 'A' and cell.row != 1: # Affect all valud cells except those in the first row or column (data cells).
            cell.fill = colourDataCell # Change background colour of data cells.
for row in worldTempWS['A24:FR24']: # Affects the row below the last data row.
    for cell in row:
        cell.border = topBorderOnly # Adds border to the bottom of the data area.
for row in worldTempWS['FS1:FS23']: # Affects the column below the last data column.
    for cell in row:
        cell.border = leftBorderOnly # Adds border to the right side od the data area.
worldTempWS['A4'].font=titleFont # Add specific font to this cell.
worldTempWS['A15'].font = titleFont # Add specific font to this cell.
print("Done.\n\n")
#End of the formatting section

###################
## Save Workbook ##
###################

print("Saving changes...")
if isfile("World Temperature.xlsx"): # If workbook already exists...
    if yesNoInput("Are you sure you would like to save changes to 'World Temperature.xlsx'? This action cannot be undone (Y/N). "): # Give user the option to save changes.
        worldTempWB.save("World Temperature.xlsx") # Save changes.
        print("\nChanges have been saved.")
    else:
        print("\nChanges have not been saved.")
else: # Workbook does not already exist.
    if yesNoInput("Would you like to save 'World Temperature.xlsx' (Y/N)? "): # Give user the option to save a new file.
        worldTempWB.save("World Temperature.xlsx") # Save new file.
        print("\nWorkbook has been saved.")
    else:
        print("\nWorkbook has not been saved.")
